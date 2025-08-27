"""
테스트 전용 픽스처 및 공통 설정
"""
import pytest
import os
import tempfile
import json
from unittest.mock import patch


@pytest.fixture
def temp_dir():
    """임시 디렉토리 픽스처"""
    with tempfile.TemporaryDirectory() as temp_dir:
        yield temp_dir


@pytest.fixture
def sample_config():
    """테스트용 설정 픽스처"""
    return {
        "repo_path": "/path/to/test/repo",
        "model_name": "qwen3:8b",
        "timeout": 600,
        "rag": {
            "enabled": True,
            "chunk_size": 1000,
            "chunk_overlap": 200
        }
    }


@pytest.fixture
def config_file(temp_dir, sample_config):
    """테스트용 config.json 파일 픽스처"""
    config_path = os.path.join(temp_dir, "config.json")
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(sample_config, f, ensure_ascii=False, indent=2)
    return config_path


@pytest.fixture
def sample_git_analysis():
    """테스트용 Git 분석 결과 픽스처"""
    return """### 커밋 메시지 목록:
- feat: 사용자 관리 기능 추가
- fix: 로그인 버그 수정

### 주요 코드 변경 내용 (diff):
--- 파일: src/user.py ---
+def create_user(name, email):
+    return User(name=name, email=email)
"""


@pytest.fixture
def sample_result_json():
    """테스트용 LLM 결과 JSON 픽스처"""
    return {
        "Scenario Description": "사용자 관리 시스템 테스트",
        "Test Scenario Name": "사용자 생성 및 로그인 테스트",
        "Test Cases": [
            {
                "ID": "TC_001",
                "절차": "1. 사용자 생성\\n2. 로그인 시도",
                "사전조건": "시스템 초기화 완료",
                "데이터": "name: test, email: test@test.com",
                "예상결과": "사용자 생성 성공\\n로그인 성공",
                "종류": "단위 테스트"
            }
        ]
    }


@pytest.fixture
def mock_excel_template(temp_dir):
    """테스트용 Excel 템플릿 픽스처"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # 템플릿 기본 구조 설정
    ws['B5'] = '개요'
    ws['F4'] = '제목'
    
    template_path = os.path.join(temp_dir, "template.xlsx")
    wb.save(template_path)
    return template_path


@pytest.fixture
def mock_git_repo(temp_dir):
    """테스트용 Git 저장소 픽스처"""
    import git
    
    repo = git.Repo.init(temp_dir)
    
    # 테스트 파일 생성
    test_file = os.path.join(temp_dir, "test.py")
    with open(test_file, 'w') as f:
        f.write("print('hello world')")
    
    repo.index.add([test_file])
    repo.index.commit("Initial commit")
    
    return repo


@pytest.fixture
def mock_ollama_response():
    """테스트용 Ollama API 응답 픽스처"""
    return {
        "response": '{"Test Cases": [{"ID": "TC_001", "절차": "테스트 절차"}]}'
    }


