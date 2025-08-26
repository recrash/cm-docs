"""
excel_writer.py 모듈 테스트
"""
import pytest
import os
import json
from unittest.mock import patch, Mock
from app.core.excel_writer import save_results_to_excel


class TestExcelWriter:
    """Excel 작성기 테스트"""
    
    def test_successful_excel_creation(self, temp_dir, mock_excel_template, sample_result_json):
        """성공적인 Excel 파일 생성 테스트"""
        # outputs 디렉토리 생성
        outputs_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(outputs_dir, exist_ok=True)
        
        with patch('src.excel_writer.datetime') as mock_datetime:
            mock_datetime.now.return_value.strftime.return_value = "20240101_120000"
            
            result_path = save_results_to_excel(sample_result_json, mock_excel_template)
            
            expected_filename = "outputs/20240101_120000_테스트_시나리오_결과.xlsx"
            assert result_path == expected_filename
            assert os.path.exists(result_path)
    
    def test_template_not_found(self, sample_result_json, capsys):
        """템플릿 파일이 없는 경우 테스트"""
        result = save_results_to_excel(sample_result_json, "nonexistent_template.xlsx")
        
        assert result is None
        captured = capsys.readouterr()
        assert "원본 템플릿 파일" in captured.out
        assert "찾을 수 없습니다" in captured.out
    
    def test_newline_conversion_in_procedure(self, temp_dir, mock_excel_template):
        """절차 필드의 개행 변환 테스트"""
        import openpyxl
        
        test_json = {
            "Scenario Description": "테스트 개요",
            "Test Scenario Name": "테스트 제목",
            "Test Cases": [
                {
                    "ID": "TC_001",
                    "절차": "1. 첫 번째 단계\\n2. 두 번째 단계\\n3. 세 번째 단계",
                    "사전조건": "초기화\\n완료",
                    "데이터": "test\\ndata",
                    "예상결과": "성공\\n완료",
                    "종류": "단위 테스트"
                }
            ]
        }
        
        outputs_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(outputs_dir, exist_ok=True)
        
        with patch('src.excel_writer.datetime') as mock_datetime:
            mock_datetime.now.return_value.strftime.return_value = "20240101_120000"
            
            result_path = save_results_to_excel(test_json, mock_excel_template)
            
            # 생성된 Excel 파일 검증
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            
            # 개행 변환 확인
            procedure_cell = sheet['B11'].value
            assert "1. 첫 번째 단계\n2. 두 번째 단계\n3. 세 번째 단계" == procedure_cell
            
            precondition_cell = sheet['C11'].value
            assert "초기화\n완료" == precondition_cell
            
            expected_result_cell = sheet['E11'].value
            assert "성공\n완료" == expected_result_cell
    
    def test_test_type_mapping(self, temp_dir, mock_excel_template):
        """테스트 종류 매핑 테스트"""
        import openpyxl
        
        test_json = {
            "Scenario Description": "테스트 개요",
            "Test Scenario Name": "테스트 제목",
            "Test Cases": [
                {
                    "ID": "TC_001",
                    "절차": "단위 테스트 절차",
                    "사전조건": "조건",
                    "데이터": "데이터",
                    "예상결과": "결과",
                    "종류": "단위 테스트"
                },
                {
                    "ID": "TC_002",
                    "절차": "통합 테스트 절차",
                    "사전조건": "조건",
                    "데이터": "데이터",
                    "예상결과": "결과",
                    "종류": "통합 테스트"
                },
                {
                    "ID": "TC_003",
                    "절차": "기타 테스트 절차",
                    "사전조건": "조건",
                    "데이터": "데이터",
                    "예상결과": "결과",
                    "종류": "기타"
                }
            ]
        }
        
        outputs_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(outputs_dir, exist_ok=True)
        
        with patch('src.excel_writer.datetime') as mock_datetime:
            mock_datetime.now.return_value.strftime.return_value = "20240101_120000"
            
            result_path = save_results_to_excel(test_json, mock_excel_template)
            
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            
            # 단위 테스트 매핑 확인
            assert sheet['F11'].value == 'Y'  # 단위 테스트
            assert sheet['G11'].value in [None, '']   # 통합 테스트 아님
            
            # 통합 테스트 매핑 확인
            assert sheet['F12'].value in [None, '']   # 단위 테스트 아님
            assert sheet['G12'].value == 'Y'  # 통합 테스트
            
            # 기타 테스트 매핑 확인
            assert sheet['F13'].value in [None, '']   # 단위 테스트 아님
            assert sheet['G13'].value in [None, '']   # 통합 테스트 아님
    
    def test_json_data_serialization(self, temp_dir, mock_excel_template):
        """JSON 데이터 직렬화 테스트"""
        import openpyxl
        
        test_json = {
            "Scenario Description": "테스트",
            "Test Scenario Name": "테스트",
            "Test Cases": [
                {
                    "ID": "TC_001",
                    "절차": "절차",
                    "사전조건": "조건",
                    "데이터": {
                        "user": "test",
                        "password": "1234",
                        "nested": {"key": "value"}
                    },
                    "예상결과": "결과",
                    "종류": "단위"
                }
            ]
        }
        
        outputs_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(outputs_dir, exist_ok=True)
        
        with patch('src.excel_writer.datetime') as mock_datetime:
            mock_datetime.now.return_value.strftime.return_value = "20240101_120000"
            
            result_path = save_results_to_excel(test_json, mock_excel_template)
            
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            
            data_cell = sheet['D11'].value
            parsed_data = json.loads(data_cell)
            
            assert parsed_data["user"] == "test"
            assert parsed_data["password"] == "1234"
            assert parsed_data["nested"]["key"] == "value"
    
    def test_empty_test_cases(self, temp_dir, mock_excel_template, capsys):
        """빈 테스트 케이스 처리 테스트"""
        test_json = {
            "Scenario Description": "테스트",
            "Test Scenario Name": "테스트",
            "Test Cases": []
        }
        
        outputs_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(outputs_dir, exist_ok=True)
        
        with patch('src.excel_writer.datetime') as mock_datetime:
            mock_datetime.now.return_value.strftime.return_value = "20240101_120000"
            
            result_path = save_results_to_excel(test_json, mock_excel_template)
            
            assert result_path is not None
            captured = capsys.readouterr()
            assert "0개의 테스트 시나리오를 저장했습니다" in captured.out
    
    def test_missing_fields_handling(self, temp_dir, mock_excel_template):
        """필수 필드 누락 처리 테스트"""
        import openpyxl
        
        test_json = {
            "Test Cases": [
                {
                    "ID": "TC_001"
                    # 다른 필드들 누락
                }
            ]
        }
        
        outputs_dir = os.path.join(temp_dir, "outputs")
        os.makedirs(outputs_dir, exist_ok=True)
        
        with patch('src.excel_writer.datetime') as mock_datetime:
            mock_datetime.now.return_value.strftime.return_value = "20240101_120000"
            
            result_path = save_results_to_excel(test_json, mock_excel_template)
            
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            
            # 기본값 확인
            assert sheet['B5'].value == "개요 생성 실패"
            assert sheet['F4'].value == "제목 생성 실패"
            assert sheet['A11'].value == "TC_001"
            assert sheet['B11'].value in [None, ""]  # 빈 문자열 또는 None