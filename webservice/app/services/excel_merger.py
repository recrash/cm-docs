"""
Excel 병합 유틸리티

Phase 2: 여러 Excel 파일을 하나로 병합하는 유틸리티
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import shutil

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side

logger = logging.getLogger(__name__)


class ExcelMergerError(Exception):
    """Excel 병합 관련 오류"""
    pass


class ExcelMerger:
    """Excel 파일 병합 클래스"""
    
    def __init__(self, documents_dir: Path):
        """
        Args:
            documents_dir: 문서 디렉토리 경로 (autodoc_service의 documents 폴더)
        """
        self.documents_dir = Path(documents_dir)
        self.merged_filename: Optional[str] = None
    
    def append_scenarios_to_base(
        self, 
        base_scenario_filename: str,
        llm_test_cases: List[Dict[str, Any]],
        session_id: str,
        change_id: str = "MERGED"
    ) -> Dict[str, Any]:
        """
        기본 시나리오 Excel 파일에 LLM이 생성한 테스트 케이스들을 추가 (성능 최적화)
        
        Args:
            base_scenario_filename: autodoc_service에서 생성한 기본 시나리오 파일명
            llm_test_cases: LLM이 생성한 테스트 케이스 목록
            session_id: 세션 ID
            change_id: 변경 ID
            
        Returns:
            병합 결과 정보
        """
        try:
            logger.info(f"기본 시나리오에 테스트 케이스 추가: {base_scenario_filename}")
            
            # 파일 존재 확인
            base_scenario_path = self.documents_dir / base_scenario_filename
            
            if not base_scenario_path.exists():
                raise ExcelMergerError(f"기본 시나리오 파일을 찾을 수 없습니다: {base_scenario_path}")
            
            # 결과 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            merged_filename = f"{timestamp}_Enhanced_Scenarios_{change_id}.xlsx"
            merged_path = self.documents_dir / merged_filename
            
            # 기본 시나리오 파일 복사
            shutil.copy2(base_scenario_path, merged_path)
            
            # 테스트 케이스 추가
            added_rows = self._append_test_cases_to_excel(merged_path, llm_test_cases)
            
            self.merged_filename = merged_filename
            
            result = {
                "success": True,
                "merged_filename": merged_filename,
                "merged_path": str(merged_path),
                "source_files": [base_scenario_filename],
                "test_cases_added": added_rows,
                "session_id": session_id
            }
            
            logger.info(f"시나리오 강화 완료: {merged_filename} ({added_rows}개 테스트 케이스 추가)")
            return result
            
        except Exception as e:
            logger.error(f"시나리오 강화 실패: {e}")
            raise ExcelMergerError(f"시나리오 강화 실패: {e}")
    
    # 기존 메서드는 하위 호환성을 위해 유지하되, 새 메서드를 호출하도록 수정
    def merge_scenario_files(
        self, 
        scenario_excel_filename: str,
        base_scenario_filename: str,
        session_id: str,
        change_id: str = "MERGED"
    ) -> Dict[str, Any]:
        """
        하위 호환성을 위한 레거시 메서드 - 새로운 방식으로 리다이렉트
        
        Args:
            scenario_excel_filename: webservice에서 생성한 시나리오 Excel 파일명 (이제 무시됨)
            base_scenario_filename: autodoc_service에서 생성한 기본 시나리오 파일명
            session_id: 세션 ID
            change_id: 변경 ID
            
        Returns:
            병합 결과 정보
        """
        # 시나리오 Excel 파일에서 테스트 케이스 추출
        try:
            scenario_path = self._find_file_in_outputs(scenario_excel_filename)
            test_cases = self._extract_test_cases_from_excel(scenario_path)
        except Exception as e:
            logger.warning(f"시나리오 파일에서 테스트 케이스 추출 실패, 빈 케이스로 처리: {e}")
            test_cases = []
        
        # 새로운 방식으로 처리
        return self.append_scenarios_to_base(
            base_scenario_filename=base_scenario_filename,
            llm_test_cases=test_cases,
            session_id=session_id,
            change_id=change_id
        )
    
    def _find_file_in_outputs(self, filename: str) -> Path:
        """
        webservice outputs 디렉토리에서 파일 찾기
        
        Args:
            filename: 찾을 파일명
            
        Returns:
            파일 경로
        """
        # webservice outputs 디렉토리에서 찾기
        from app.core.paths import get_outputs_dir
        outputs_dir = get_outputs_dir()
        
        file_path = outputs_dir / filename
        if file_path.exists():
            return file_path
        
        # 패턴 매칭으로 찾기 (timestamp가 다를 수 있음)
        for file in outputs_dir.glob("*.xlsx"):
            if filename.split("_")[-1] in file.name:  # 파일명 끝부분 매칭
                return file
        
        # 기본 경로 반환 (존재하지 않을 수 있음)
        return outputs_dir / filename
    
    def _append_test_cases_to_excel(self, excel_path: Path, test_cases: List[Dict[str, Any]]) -> int:
        """
        Excel 파일에 테스트 케이스 행들을 추가
        
        Args:
            excel_path: 기본 시나리오 Excel 파일 경로
            test_cases: 추가할 테스트 케이스 목록
            
        Returns:
            추가된 행 수
        """
        if not test_cases:
            logger.info("추가할 테스트 케이스가 없습니다.")
            return 0
        
        wb = openpyxl.load_workbook(excel_path)
        
        # 첫 번째 워크시트를 기본으로 사용 (보통 테스트 시나리오 시트)
        ws = wb.active
        if not ws:
            ws = wb.worksheets[0]
        
        # 마지막 데이터 행 찾기
        last_row = self._find_last_data_row(ws)
        next_row = last_row + 1
        
        added_rows = 0
        
        # 테스트 케이스를 행으로 추가
        for i, test_case in enumerate(test_cases):
            current_row = next_row + i
            
            # 기본 테스트 케이스 구조에 맞춰 셀 채우기
            # ID (A열)
            test_id = test_case.get("ID", f"TC_{current_row:03d}")
            ws[f"A{current_row}"] = test_id
            
            # 절차 (B열)
            procedure = test_case.get("절차", test_case.get("procedure", ""))
            ws[f"B{current_row}"] = procedure
            
            # 사전조건 (C열)
            precondition = test_case.get("사전조건", test_case.get("precondition", ""))
            ws[f"C{current_row}"] = precondition
            
            # 데이터 (D열) - JSON이면 문자열로 변환
            test_data = test_case.get("데이터", test_case.get("data", ""))
            if isinstance(test_data, (dict, list)):
                import json
                test_data = json.dumps(test_data, ensure_ascii=False, indent=2)
            ws[f"D{current_row}"] = test_data
            
            # 예상결과 (E열)
            expected = test_case.get("예상결과", test_case.get("expected", ""))
            ws[f"E{current_row}"] = expected
            
            # Unit 테스트 (F열)
            unit_flag = test_case.get("Unit", "")
            ws[f"F{current_row}"] = unit_flag
            
            # Integration 테스트 (G열)  
            integration_flag = test_case.get("Integration", "")
            ws[f"G{current_row}"] = integration_flag
            
            # 스타일 적용 (기본 폰트)
            for col in range(1, 8):  # A-G 컬럼
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(name="맑은 고딕", size=10)
                
            added_rows += 1
        
        # 파일 저장
        wb.save(excel_path)
        wb.close()
        
        logger.info(f"테스트 케이스 {added_rows}개 추가 완료")
        return added_rows
    
    def _find_last_data_row(self, worksheet: Worksheet) -> int:
        """
        워크시트에서 마지막 데이터 행을 찾기
        
        Args:
            worksheet: 대상 워크시트
            
        Returns:
            마지막 데이터 행 번호 (1-based)
        """
        # A열을 기준으로 마지막 데이터 찾기
        last_row = 1
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet[f"A{row}"].value
            if cell_value is not None and str(cell_value).strip():
                last_row = row
        
        return last_row
    
    def _extract_test_cases_from_excel(self, excel_path: Path) -> List[Dict[str, Any]]:
        """
        Excel 파일에서 테스트 케이스 추출 (하위 호환성용)
        
        Args:
            excel_path: Excel 파일 경로
            
        Returns:
            테스트 케이스 목록
        """
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            test_cases = []
            
            # 헤더 행 스킵하고 데이터 행부터 읽기 (보통 11행부터 시작)
            start_row = 11
            
            for row in range(start_row, ws.max_row + 1):
                # A열이 비어있으면 종료
                if not ws[f"A{row}"].value:
                    break
                
                test_case = {
                    "ID": ws[f"A{row}"].value or f"TC_{row:03d}",
                    "절차": ws[f"B{row}"].value or "",
                    "사전조건": ws[f"C{row}"].value or "",
                    "데이터": ws[f"D{row}"].value or "",
                    "예상결과": ws[f"E{row}"].value or "",
                    "Unit": ws[f"F{row}"].value or "",
                    "Integration": ws[f"G{row}"].value or ""
                }
                
                test_cases.append(test_case)
            
            wb.close()
            return test_cases
            
        except Exception as e:
            logger.error(f"Excel에서 테스트 케이스 추출 실패: {e}")
            return []
    
    def get_merged_file_info(self) -> Optional[Dict[str, Any]]:
        """
        병합된 파일 정보 반환
        
        Returns:
            파일 정보 딕셔너리 또는 None
        """
        if not self.merged_filename:
            return None
        
        merged_path = self.documents_dir / self.merged_filename
        if not merged_path.exists():
            return None
        
        return {
            "filename": self.merged_filename,
            "path": str(merged_path),
            "size": merged_path.stat().st_size,
            "created_at": datetime.fromtimestamp(merged_path.stat().st_ctime).isoformat()
        }


def test_excel_merger():
    """Excel 병합 테스트 함수"""
    try:
        # 테스트용 임시 디렉토리
        test_dir = Path("/tmp/excel_merger_test")
        test_dir.mkdir(exist_ok=True)
        
        merger = ExcelMerger(test_dir)
        
        # 테스트 파일 생성 (실제로는 autodoc_service와 webservice에서 생성됨)
        print("Excel 병합 유틸리티 테스트 준비 완료")
        print(f"테스트 디렉토리: {test_dir}")
        
    except Exception as e:
        print(f"Excel 병합 테스트 실패: {e}")


if __name__ == "__main__":
    test_excel_merger()