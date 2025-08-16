"""
Excel 테스트 시나리오 빌더 테스트

openpyxl로 생성된 Excel 파일의 셀 좌표 매핑 검증
"""
import json
import pytest
import tempfile
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

from ..models import ChangeRequest, ChangeDetails
from ..services.excel_test_builder import build_test_scenario_xlsx


class TestExcelTestBuilder:
    """Excel 테스트 시나리오 빌더 테스트"""
    
    @pytest.fixture
    def sample_change_request(self):
        """테스트용 변경 요청 데이터"""
        fixtures_dir = Path(__file__).parent / "fixtures"
        json_file = fixtures_dir / "itsupp_sample.json"
        
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return ChangeRequest(
            change_id=data["변경관리번호"],
            system=data["시스템"],
            system_short=data["시스템_약칭"],
            title=data["제목"],
            writer_short=data["처리자_약칭"],
            biz_test_date=data.get("현업_테스트_일자", None)
        )
    
    @pytest.fixture
    def temp_output_dir(self):
        """임시 출력 디렉터리"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)
    
    def test_build_excel_file_success(self, sample_change_request, temp_output_dir):
        """Excel 파일 생성 성공 테스트"""
        output_path = build_test_scenario_xlsx(sample_change_request, temp_output_dir)
        
        # 파일이 생성되었는지 확인
        assert output_path.exists(), "Excel 파일이 생성되지 않았습니다"
        assert output_path.suffix == '.xlsx', "파일 확장자가 .xlsx가 아닙니다"
        
        # 파일 크기 확인
        assert output_path.stat().st_size > 0, "생성된 Excel 파일이 비어있습니다"
    
    def test_excel_filename_format(self, sample_change_request, temp_output_dir):
        """Excel 파일명 형식 검증"""
        output_path = build_test_scenario_xlsx(sample_change_request, temp_output_dir)
        
        filename = output_path.name
        
        # 파일명 형식: [YYMMDD 처리자] 테스트시나리오 {change_id} {title}.xlsx
        assert filename.startswith('['), "파일명이 대괄호로 시작하지 않습니다"
        assert '테스트시나리오' in filename, "파일명에 '테스트시나리오'가 없습니다"
        assert sample_change_request.change_id in filename, "파일명에 change_id가 없습니다"
        assert filename.endswith('.xlsx'), "파일명이 .xlsx로 끝나지 않습니다"
    
    def test_excel_cell_mapping_basic(self, sample_change_request, temp_output_dir):
        """Excel 기본 셀 매핑 검증"""
        output_path = build_test_scenario_xlsx(sample_change_request, temp_output_dir)
        
        # Excel 파일 열기
        wb = load_workbook(str(output_path))
        
        # 첫 번째 워크시트 또는 '테스트' 시트
        if '테스트' in wb.sheetnames:
            ws = wb['테스트']
        else:
            ws = wb.active
        
        # 기본 셀 매핑 검증
        assert ws['C2'].value == sample_change_request.system, (
            f"C2 셀 값이 일치하지 않습니다. 실제: '{ws['C2'].value}', 기대: '{sample_change_request.system}'"
        )
        
        assert ws['F2'].value == sample_change_request.change_id, (
            f"F2 셀 값이 일치하지 않습니다. 실제: '{ws['F2'].value}', 기대: '{sample_change_request.change_id}'"
        )
        
        assert ws['B4'].value == sample_change_request.system, (
            f"B4 셀 값이 일치하지 않습니다. 실제: '{ws['B4'].value}', 기대: '{sample_change_request.system}'"
        )
        
        assert ws['D4'].value == sample_change_request.change_id, (
            f"D4 셀 값이 일치하지 않습니다. 실제: '{ws['D4'].value}', 기대: '{sample_change_request.change_id}'"
        )
        
        assert ws['F4'].value == sample_change_request.title, (
            f"F4 셀 값이 일치하지 않습니다. 실제: '{ws['F4'].value}', 기대: '{sample_change_request.title}'"
        )
        
        wb.close()
    
    def test_biz_test_date_with_value(self, temp_output_dir):
        """현업 테스트 일자가 있는 경우 테스트"""
        request_with_test_date = ChangeRequest(
            change_id="TEST_001",
            system="테스트시스템",
            title="테스트 제목",
            biz_test_date="2025-08-15"
        )
        
        output_path = build_test_scenario_xlsx(request_with_test_date, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb['테스트'] if '테스트' in wb.sheetnames else wb.active
        
        # biz_test_date가 있으면 B7에 해당 값, K7에 2
        assert ws['B7'].value == "2025-08-15", (
            f"B7 셀 값이 일치하지 않습니다. 실제: '{ws['B7'].value}', 기대: '2025-08-15'"
        )
        
        assert ws['K7'].value == 2, (
            f"K7 셀 값이 일치하지 않습니다. 실제: '{ws['K7'].value}', 기대: 2"
        )
        
        wb.close()
    
    def test_biz_test_date_without_value(self, temp_output_dir):
        """현업 테스트 일자가 없는 경우 테스트"""
        request_without_test_date = ChangeRequest(
            change_id="TEST_001",
            system="테스트시스템",
            title="테스트 제목",
            biz_test_date=None
        )
        
        output_path = build_test_scenario_xlsx(request_without_test_date, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb['테스트'] if '테스트' in wb.sheetnames else wb.active
        
        # biz_test_date가 없으면 B7에 오늘 날짜, K7에 1
        today = datetime.now().strftime("%Y-%m-%d")
        
        assert ws['B7'].value == today, (
            f"B7 셀 값이 일치하지 않습니다. 실제: '{ws['B7'].value}', 기대: '{today}'"
        )
        
        assert ws['K7'].value == 1, (
            f"K7 셀 값이 일치하지 않습니다. 실제: '{ws['K7'].value}', 기대: 1"
        )
        
        wb.close()
    
    def test_all_cell_mappings(self, sample_change_request, temp_output_dir):
        """모든 셀 매핑 검증"""
        output_path = build_test_scenario_xlsx(sample_change_request, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb['테스트'] if '테스트' in wb.sheetnames else wb.active
        
        # 모든 매핑 검증
        expected_mappings = {
            'C2': sample_change_request.system,
            'F2': sample_change_request.change_id,
            'B4': sample_change_request.system,
            'D4': sample_change_request.change_id,
            'F4': sample_change_request.title,
            'B5': sample_change_request.title,
            'F7': sample_change_request.writer_short or "",
            'A11': sample_change_request.system,
            'B11': sample_change_request.title
        }
        
        for cell_ref, expected_value in expected_mappings.items():
            actual_value = ws[cell_ref].value
            assert actual_value == expected_value, (
                f"{cell_ref} 셀 값이 일치하지 않습니다. "
                f"실제: '{actual_value}', 기대: '{expected_value}'"
            )
        
        wb.close()
    
    def test_missing_required_fields_error(self, temp_output_dir):
        """필수 필드 누락 시 에러 발생 테스트"""
        # change_id 누락
        with pytest.raises(ValueError, match="change_id는 필수입니다"):
            build_test_scenario_xlsx(
                ChangeRequest(change_id="", system="테스트", title="테스트"),
                temp_output_dir
            )
        
        # title 누락  
        with pytest.raises(ValueError, match="title은 필수입니다"):
            build_test_scenario_xlsx(
                ChangeRequest(change_id="TEST_001", system="테스트", title=""),
                temp_output_dir
            )
        
        # system 누락
        with pytest.raises(ValueError, match="system은 필수입니다"):
            build_test_scenario_xlsx(
                ChangeRequest(change_id="TEST_001", system="", title="테스트"),
                temp_output_dir
            )
    
    def test_missing_template_error(self, sample_change_request, temp_output_dir, monkeypatch):
        """템플릿 파일 누락 시 에러 발생 테스트"""
        def mock_verify_template_exists(template_name):
            raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_name}")
        
        monkeypatch.setattr(
            "autodoc_service.app.services.excel_test_builder.verify_template_exists",
            mock_verify_template_exists
        )
        
        with pytest.raises(FileNotFoundError, match="템플릿 파일을 찾을 수 없습니다"):
            build_test_scenario_xlsx(sample_change_request, temp_output_dir)
    
    def test_output_directory_creation(self, sample_change_request):
        """출력 디렉터리가 없으면 기본 디렉터리 사용"""
        # out_dir을 None으로 전달하면 기본 documents 디렉터리 사용
        output_path = build_test_scenario_xlsx(sample_change_request, None)
        
        # 파일이 생성되었는지만 확인 (실제 documents 디렉터리에 생성됨)
        assert output_path.exists(), "기본 디렉터리에 파일이 생성되지 않았습니다"
        assert output_path.suffix == '.xlsx', "파일 확장자가 올바르지 않습니다"
        
        # 테스트 후 정리
        if output_path.exists():
            output_path.unlink()
    
    def test_writer_short_empty_handling(self, temp_output_dir):
        """writer_short가 없는 경우 처리 테스트"""
        request_without_writer = ChangeRequest(
            change_id="TEST_001",
            system="테스트시스템",
            title="테스트 제목",
            writer_short=None
        )
        
        output_path = build_test_scenario_xlsx(request_without_writer, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb['테스트'] if '테스트' in wb.sheetnames else wb.active
        
        # F7 셀이 빈 문자열이어야 함
        assert ws['F7'].value == "", (
            f"F7 셀이 빈 값이어야 합니다. 실제: '{ws['F7'].value}'"
        )
        
        wb.close()