"""
Excel 변경관리 목록 빌더 테스트

openpyxl로 생성된 Excel 목록 파일의 테이블 append 검증
"""
import pytest
import tempfile
from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import load_workbook

from ..models import ChangeRequest, ChangeDetails
from ..services.excel_list_builder import build_change_list_xlsx


class TestExcelListBuilder:
    """Excel 변경관리 목록 빌더 테스트"""
    
    @pytest.fixture
    def sample_change_requests(self):
        """테스트용 변경 요청 데이터 목록"""
        return [
            ChangeRequest(
                change_id="LIMS_20250814_1",
                system="LIMS-001",
                system_short="LIMS",
                title="첫 번째 변경사항",
                requester="홍길동",
                deploy_datetime="08/15 13:00",
                biz_test_date="2025-08-14",
                deployer="김배포"
            ),
            ChangeRequest(
                change_id="OLIVE_20250814_2", 
                system="OLIVE-002",
                system_short="OLIVE",
                title="두 번째 변경사항",
                requester="이순신",
                deploy_datetime=None,  # 없으면 내일 날짜
                biz_test_date=None,
                deployer="박배포"
            )
        ]
    
    @pytest.fixture
    def temp_output_dir(self):
        """임시 출력 디렉터리"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)
    
    def test_build_excel_list_success(self, sample_change_requests, temp_output_dir):
        """Excel 목록 파일 생성 성공 테스트"""
        output_path = build_change_list_xlsx(sample_change_requests, temp_output_dir)
        
        # 파일이 생성되었는지 확인
        assert output_path.exists(), "Excel 목록 파일이 생성되지 않았습니다"
        assert output_path.suffix == '.xlsx', "파일 확장자가 .xlsx가 아닙니다"
        
        # 파일 크기 확인
        assert output_path.stat().st_size > 0, "생성된 Excel 파일이 비어있습니다"
    
    def test_excel_list_filename_format(self, sample_change_requests, temp_output_dir):
        """Excel 목록 파일명 형식 검증"""
        output_path = build_change_list_xlsx(sample_change_requests, temp_output_dir)
        
        filename = output_path.name
        today = datetime.now().strftime("%Y%m%d")
        
        # 파일명 형식: 변경관리목록_{YYYYMMDD}.xlsx
        expected_filename = f"변경관리목록_{today}.xlsx"
        assert filename == expected_filename, (
            f"파일명이 일치하지 않습니다. 실제: '{filename}', 기대: '{expected_filename}'"
        )
    
    def test_excel_list_column_order(self, sample_change_requests, temp_output_dir):
        """Excel 목록 열 순서 검증"""
        output_path = build_change_list_xlsx(sample_change_requests, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        try:
            ws = wb.active
            
            # '변경관리번호'(10열)로 실제 데이터 행 찾기
            first_item = sample_change_requests[0]
            row = None
            for row_num in range(1, ws.max_row + 1):
                if ws.cell(row_num, 10).value == first_item.change_id:
                    row = row_num
                    break

            assert row is not None, "변경관리번호로 데이터 행을 찾을 수 없습니다"
            
            # 11열 순서 검증
            expected_values = [
                first_item.deploy_type or "정기배포",        # 1) 배포종류
                first_item.system_short or first_item.system,  # 2) 시스템
                first_item.biz_test_date or "",               # 3) 현업 테스트 일자
                "8월 15일",                                   # 4) 배포일자 (08/15 13:00 → 8월 15일)
                first_item.requester or "",                   # 5) 요청자
                "",                                           # 6) IT 지원의뢰서 (없음)
                first_item.program or "Appl.",               # 7) Program
                "",                                          # 8) 소스명 (외부 입력)
                first_item.deployer or "",                   # 9) 배포자
                first_item.change_id,                        # 10) 변경관리번호
                first_item.has_cm_doc or "O"                 # 11) 변경관리문서유무
            ]
            
            for col, expected_value in enumerate(expected_values, 1):
                actual_value = ws.cell(row, col).value
                if expected_value == "":
                    assert actual_value in ("", None), (
                        f"행 {row}, 열 {col} 값이 일치하지 않습니다. 실제: '{actual_value}', 기대: '' 또는 None"
                    )
                else:
                    assert actual_value == expected_value, (
                        f"행 {row}, 열 {col} 값이 일치하지 않습니다. "
                        f"실제: '{actual_value}', 기대: '{expected_value}'"
                    )
        finally:
            wb.close()
    
    def test_deploy_date_formatting_with_datetime(self, temp_output_dir):
        """deploy_datetime가 있는 경우 배포일자 포맷팅 테스트"""
        request_with_deploy_time = ChangeRequest(
            change_id="TEST_001",
            system="테스트",
            title="배포일시 테스트",
            deploy_datetime="08/20 14:30"
        )
        
        output_path = build_change_list_xlsx([request_with_deploy_time], temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb.active
        
        # 변경관리번호로 데이터 행 찾기 후 4열 확인
        target_row = None
        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row_num, 10).value == "TEST_001":
                target_row = row_num
                break
        assert target_row is not None, "데이터 행을 찾지 못했습니다"
        deploy_date = ws.cell(target_row, 4).value
        assert deploy_date == "8월 20일", (
            f"배포일자 포맷이 올바르지 않습니다. 실제: '{deploy_date}', 기대: '8월 20일'"
        )
        
        wb.close()
    
    def test_deploy_date_formatting_without_datetime(self, temp_output_dir):
        """deploy_datetime가 없는 경우 내일 날짜 사용 테스트"""
        request_without_deploy_time = ChangeRequest(
            change_id="TEST_001",
            system="테스트",
            title="배포일시 없음 테스트",
            deploy_datetime=None
        )
        
        output_path = build_change_list_xlsx([request_without_deploy_time], temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb.active
        
        # 내일 날짜 계산
        tomorrow = datetime.now() + timedelta(days=1)
        expected_date = f"{tomorrow.month}월 {tomorrow.day}일"
        
        # 변경관리번호로 데이터 행 찾기 후 4열 확인
        target_row = None
        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row_num, 10).value == "TEST_001":
                target_row = row_num
                break
        assert target_row is not None, "데이터 행을 찾지 못했습니다"
        deploy_date = ws.cell(target_row, 4).value
        assert deploy_date == expected_date, (
            f"배포일자 포맷이 올바르지 않습니다. 실제: '{deploy_date}', 기대: '{expected_date}'"
        )
        
        wb.close()
    
    def test_multiple_items_append(self, sample_change_requests, temp_output_dir):
        """여러 항목 append 테스트"""
        output_path = build_change_list_xlsx(sample_change_requests, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb.active
        
        # 입력한 변경관리번호들의 존재 여부 확인
        expected_ids = [item.change_id for item in sample_change_requests]
        found_ids = []
        for row_num in range(1, ws.max_row + 1):
            cid = ws.cell(row_num, 10).value
            if cid in expected_ids:
                found_ids.append(cid)

        assert sorted(found_ids) == sorted(expected_ids), (
            f"변경관리번호 목록이 일치하지 않습니다. 실제: {found_ids}, 기대: {expected_ids}"
        )
        
        wb.close()
    
    def test_dict_input_handling(self, temp_output_dir):
        """딕셔너리 입력 처리 테스트"""
        dict_items = [
            {
                'deploy_type': '긴급배포',
                'system': 'TEST-001',
                'biz_test_date': '2025-08-14',
                'deploy_datetime': '08/14 15:00',
                'requester': '테스트사용자',
                'program': 'Web',
                'deployer': '테스트배포자',
                'change_id': 'DICT_TEST_001',
                'has_cm_doc': 'X'
            }
        ]
        
        output_path = build_change_list_xlsx(dict_items, temp_output_dir)
        
        wb = load_workbook(str(output_path))
        ws = wb.active
        
        # 변경관리번호로 데이터 행 찾기
        target_row = None
        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row_num, 10).value == 'DICT_TEST_001':
                target_row = row_num
                break
        assert target_row is not None, "데이터 행을 찾지 못했습니다"
        assert ws.cell(target_row, 1).value == '긴급배포', "배포종류가 일치하지 않습니다"
        assert ws.cell(target_row, 7).value == 'Web', "Program이 일치하지 않습니다"
        assert ws.cell(target_row, 10).value == 'DICT_TEST_001', "변경관리번호가 일치하지 않습니다"
        assert ws.cell(target_row, 11).value == 'X', "변경관리문서유무가 일치하지 않습니다"
        
        wb.close()
    
    def test_empty_items_error(self, temp_output_dir):
        """빈 항목 리스트 에러 테스트"""
        with pytest.raises(ValueError, match="items는 비어있을 수 없습니다"):
            build_change_list_xlsx([], temp_output_dir)
    
    def test_missing_template_error(self, sample_change_requests, temp_output_dir, monkeypatch):
        """템플릿 파일 누락 시 에러 발생 테스트 (환경변수로 빈 템플릿 폴더 지정)"""
        empty_dir = temp_output_dir / "empty_tpl"
        empty_dir.mkdir(parents=True, exist_ok=True)
        monkeypatch.setenv("AUTODOC_TPL_DIR", str(empty_dir))
        
        with pytest.raises(FileNotFoundError, match="템플릿 파일을 찾을 수 없습니다"):
            build_change_list_xlsx(sample_change_requests, temp_output_dir)
    
    def test_filename_sanitization(self, temp_output_dir):
        """파일명 sanitize 테스트 (중복 방지)"""
        simple_request = [ChangeRequest(
            change_id="TEST_001",
            system="테스트",
            title="중복 테스트"
        )]
        
        # 같은 파일을 두 번 생성
        output_path1 = build_change_list_xlsx(simple_request, temp_output_dir)
        output_path2 = build_change_list_xlsx(simple_request, temp_output_dir)
        
        # 두 파일의 경로가 달라야 함 (중복 방지)
        assert output_path1 != output_path2, "중복 파일명이 생성되었습니다"
        assert output_path1.exists() and output_path2.exists(), "두 파일 모두 생성되어야 합니다"
        
        # 파일명 패턴 확인
        today = datetime.now().strftime("%Y%m%d")
        base_name = f"변경관리목록_{today}.xlsx"
        
        assert output_path1.name == base_name, f"첫 번째 파일명이 올바르지 않습니다: {output_path1.name}"
        assert "_1" in output_path2.name, f"두 번째 파일명에 증분 번호가 없습니다: {output_path2.name}"