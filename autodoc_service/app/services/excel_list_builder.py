"""
Excel 변경관리 문서 목록 생성 서비스

openpyxl을 사용하여 템플릿 기반 Excel 목록 생성
테이블 append 방식으로 데이터 추가
"""
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Union, Dict, Any

from openpyxl import load_workbook

from ..models import ChangeRequest
from .paths import verify_template_exists, get_documents_dir
from .filename import generate_excel_list_filename, unique_path


def _format_deploy_date(deploy_datetime: Optional[str]) -> str:
    """배포일자 포맷팅
    
    deploy_datetime가 있으면 날짜만 추출, 없으면 내일 날짜를 "M월 d일" 형식으로 반환
    """
    if deploy_datetime:
        # "08/07 13:00" 형식에서 날짜 부분만 추출
        if ' ' in deploy_datetime:
            date_part = deploy_datetime.split(' ')[0]
            # "08/07" → "8월 7일"
            try:
                month, day = date_part.split('/')
                return f"{int(month)}월 {int(day)}일"
            except:
                return deploy_datetime
        else:
            return deploy_datetime
    else:
        # 내일 날짜를 "M월 d일" 형식으로
        tomorrow = datetime.now() + timedelta(days=1)
        return f"{tomorrow.month}월 {tomorrow.day}일"


def _extract_filename_only(file_path: Optional[str]) -> str:
    """파일 경로에서 파일명만 추출"""
    if not file_path:
        return ""
    
    # 경로에서 파일명만 추출
    if '/' in file_path or '\\' in file_path:
        return Path(file_path).name
    
    return file_path


def build_change_list_xlsx(
    items: List[Union[ChangeRequest, Dict[str, Any]]], 
    out_dir: Optional[Path] = None
) -> Path:
    """변경관리 문서 목록 Excel 파일 생성
    
    Args:
        items: 변경관리 요청 데이터 목록
        out_dir: 출력 디렉터리 (None이면 기본 documents 디렉터리 사용)
        
    Returns:
        Path: 생성된 파일 경로
        
    Raises:
        FileNotFoundError: 템플릿 파일이 없는 경우
        ValueError: 데이터가 없는 경우
    """
    if not items:
        raise ValueError("items는 비어있을 수 없습니다")
    
    # 템플릿 로드
    template_path = verify_template_exists("template_list.xlsx")
    wb = load_workbook(str(template_path))
    
    # 출력 디렉터리 설정
    if out_dir is None:
        out_dir = get_documents_dir()
    
    # 첫 번째 워크시트 사용
    ws = wb.active
    
    # 기존 데이터가 있는 마지막 행 찾기
    max_row = ws.max_row
    
    # 헤더가 있다고 가정하고 데이터 행부터 시작
    start_row = max_row + 1
    
    # 각 항목을 행으로 추가
    for i, item in enumerate(items):
        # ChangeRequest 객체를 dict로 변환
        if isinstance(item, ChangeRequest):
            data = {
                'deploy_type': item.deploy_type or "정기배포",
                'system': item.system_short or item.system,
                'biz_test_date': item.biz_test_date or "",
                'deploy_datetime': item.deploy_datetime,
                'requester': item.requester or "",
                'it_request_html': item.it_request_html or "",
                'program': item.program or "Appl.",
                'deployer': item.deployer or "",
                'change_id': item.change_id,
                'has_cm_doc': item.has_cm_doc or "O"
            }
        else:
            data = item
        
        row_num = start_row + i
        
        # 11열 데이터 매핑
        columns = [
            data.get('deploy_type', '정기배포'),                    # 1) 배포종류
            data.get('system', ''),                                # 2) 시스템
            data.get('biz_test_date', ''),                         # 3) 현업 테스트 일자
            _format_deploy_date(data.get('deploy_datetime')),       # 4) 배포일자
            data.get('requester', ''),                             # 5) 요청자
            _extract_filename_only(data.get('it_request_html')),   # 6) IT 지원의뢰서
            data.get('program', 'Appl.'),                          # 7) Program
            data.get('source_name', ''),                           # 8) 소스명 (외부 입력)
            data.get('deployer', ''),                              # 9) 배포자
            data.get('change_id', ''),                             # 10) 변경관리번호
            data.get('has_cm_doc', 'O')                            # 11) 변경관리문서유무
        ]
        
        # 각 열에 데이터 입력 (A, B, C, ... K 열)
        for col_idx, value in enumerate(columns, 1):
            ws.cell(row=row_num, column=col_idx, value=value)
    
    # 파일명 생성
    filename = generate_excel_list_filename()
    
    # 중복 방지 경로 생성
    output_path = unique_path(out_dir, filename)
    
    # 파일 저장
    wb.save(str(output_path))
    wb.close()
    
    return output_path