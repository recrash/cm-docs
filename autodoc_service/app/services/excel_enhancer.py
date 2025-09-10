"""
Excel 시나리오 강화 유틸리티

기본 시나리오 Excel에 LLM 테스트 케이스를 행 추가하는 기능
"""
import logging
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

logger = logging.getLogger(__name__)


def append_test_cases_to_excel(excel_path: Path, test_cases: List[Dict[str, Any]]) -> int:
    """
    Excel 파일에 테스트 케이스 행들을 추가
    
    Args:
        excel_path: 기본 시나리오 Excel 파일 경로
        test_cases: 추가할 테스트 케이스 목록
        
    Returns:
        추가된 행 수
    """
    if not test_cases:
        logger.info("추가할 테스트 케이스가 없습니다.")
        return 0
    
    logger.info(f"Excel 파일에 테스트 케이스 추가: {len(test_cases)}개")
    
    wb = openpyxl.load_workbook(excel_path)
    
    # 첫 번째 워크시트를 기본으로 사용 (보통 테스트 시나리오 시트)
    ws = wb.active
    if not ws:
        ws = wb.worksheets[0]
    
    # 11행부터 시작 (사용자 요구사항)
    start_row = 11
    
    # 스타일 정의
    thin_border = Side(style='thin', color='000000')
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    font = Font(name="맑은 고딕", size=10)
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    added_rows = 0
    
    # 1단계: 모든 테스트 케이스 데이터를 먼저 입력
    for i, test_case in enumerate(test_cases):
        current_row = start_row + i
        
        # 기본 테스트 케이스 구조에 맞춰 셀 채우기
        # ID (A열)
        test_id = test_case.get("ID", f"TC_{current_row:03d}")
        ws[f"A{current_row}"] = test_id
        
        # 절차 (B열)
        procedure = test_case.get("절차", test_case.get("procedure", ""))
        ws[f"B{current_row}"] = procedure
        
        # 사전조건 (C열)
        precondition = test_case.get("사전조건", test_case.get("precondition", ""))
        ws[f"C{current_row}"] = precondition
        
        # 데이터 (D열) - JSON이면 문자열로 변환
        test_data = test_case.get("데이터", test_case.get("data", ""))
        if isinstance(test_data, (dict, list)):
            import json
            test_data = json.dumps(test_data, ensure_ascii=False, indent=2)
        ws[f"D{current_row}"] = test_data
        
        # 예상결과 (E열)
        expected = test_case.get("예상결과", test_case.get("expected", ""))
        ws[f"E{current_row}"] = expected
        
        # Unit 테스트 (F열)
        unit_flag = test_case.get("Unit", "")
        ws[f"F{current_row}"] = unit_flag
        
        # Integration 테스트 (G열)  
        integration_flag = test_case.get("Integration", "")
        ws[f"G{current_row}"] = integration_flag
        
        added_rows += 1
    
    # 2단계: 모든 데이터 입력 완료 후 스타일 적용
    header_row = 10
    end_row = start_row + added_rows - 1
    
    # A10~K열 범위에 모든 스타일 적용
    for row in range(header_row, end_row + 1):
        # 행 높이 고정 (데이터 행만)
        if row >= start_row:
            ws.row_dimensions[row].height = 80
        
        # 모든 열에 스타일 적용
        for col in range(1, 12):  # A-K 컬럼
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            # 데이터 행에만 폰트, 정렬 적용
            if row >= start_row:
                cell.font = font
                cell.alignment = alignment
    
    # 파일 저장
    wb.save(excel_path)
    wb.close()
    
    logger.info(f"테스트 케이스 {added_rows}개 추가 완료")
    return added_rows


def _find_last_data_row(worksheet) -> int:
    """
    워크시트에서 마지막 데이터 행을 찾기
    
    Args:
        worksheet: 대상 워크시트
        
    Returns:
        마지막 데이터 행 번호 (1-based)
    """
    # A열을 기준으로 마지막 데이터 찾기
    last_row = 1
    for row in range(1, worksheet.max_row + 1):
        cell_value = worksheet[f"A{row}"].value
        if cell_value is not None and str(cell_value).strip():
            last_row = row
    
    return last_row