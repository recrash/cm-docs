"""
## B) 테스트 시나리오 (template.xlsx)
- 템플릿: templates/template.xlsx (로드만)
- 시트명: `테스트` (존재 가정)
- 셀 매핑(고정 좌표):
  - C2 ← system
  - F2 ← change_id
  - B4 ← system
  - D4 ← change_id
  - F4 ← title
  - B5 ← title
  - B7 ← biz_test_date가 있으면 그 값(YYYY-MM-DD), 없으면 오늘
  - K7 ← biz_test_date 있으면 2, 없으면 1
  - F7 ← writer_short
  - A11 ← system
  - B11 ← title
- 출력 파일명: `[YYMMDD 처리자] 테스트시나리오 {change_id} {title}.docx`
- 동일 sanitize 적용
- ***인덱스 및 좌표 임의 변경 금지***
"""
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font

from ..models import ChangeRequest
from .paths import verify_template_exists, get_documents_dir
from .filename import generate_excel_test_filename, unique_path


def build_test_scenario_xlsx(data: ChangeRequest, out_dir: Optional[Path] = None) -> Path:
    """테스트 시나리오 Excel 파일 생성
    
    Args:
        data: 변경 요청 데이터
        out_dir: 출력 디렉터리 (None이면 기본 documents 디렉터리 사용)
        
    Returns:
        Path: 생성된 파일 경로
        
    Raises:
        FileNotFoundError: 템플릿 파일이 없는 경우
        ValueError: 필수 데이터가 없는 경우
    """
    # 필수 데이터 검증
    if not data.change_id:
        raise ValueError("change_id는 필수입니다")
    if not data.title:
        raise ValueError("title은 필수입니다")
    if not data.system:
        raise ValueError("system은 필수입니다")
    
    # 템플릿 로드
    template_path = verify_template_exists("template.xlsx")
    wb = load_workbook(str(template_path))
    
    # 출력 디렉터리 설정
    if out_dir is None:
        out_dir = get_documents_dir()
    
    # '테스트' 시트 선택
    try:
        ws = wb['테스트']
    except KeyError:
        # 첫 번째 시트 사용 (템플릿에 '테스트' 시트가 없는 경우)
        ws = wb.active
    
    # 현업 테스트 일자 처리
    biz_test_date = data.biz_test_date
    if not biz_test_date:
        biz_test_date = datetime.now().strftime("%Y-%m-%d")
        k7_value = 1  # 없으면 1
    else:
        k7_value = 2  # 있으면 2
    
    # 셀 매핑 (고정 좌표)
    cell_mappings = {
        'C2': data.system,                    # system
        'F2': data.change_id,                 # change_id
        'B4': data.system,                    # system
        'D4': data.change_id,                 # change_id
        'F4': data.title,                     # title
        'B5': data.title,                     # title
        'B7': biz_test_date,                  # biz_test_date (있으면 그 값, 없으면 오늘)
        'K7': k7_value,                       # biz_test_date 유무에 따른 값
        'F7': data.writer_short or "",        # writer_short
        'A11': data.system,                   # system
        'B11': data.title                     # title
    }
    
    # 매핑 적용 (맑은 고딕 폰트와 함께)
    malgun_gothic_font = Font(name='맑은 고딕')
    
    for cell_ref, value in cell_mappings.items():
        try:
            cell = ws[cell_ref]
            cell.value = value
            # 맑은 고딕 폰트 적용
            cell.font = malgun_gothic_font
        except Exception as e:
            # 셀 참조 오류는 로그만 남기고 계속 진행
            print(f"Warning: 셀 {cell_ref} 매핑 실패: {e}")
    
    # 파일명 생성
    filename = generate_excel_test_filename(
        change_id=data.change_id,
        title=data.title,
        writer_short=data.writer_short
    )
    
    # 중복 방지 경로 생성
    output_path = unique_path(out_dir, filename)
    
    # 파일 저장
    wb.save(str(output_path))
    wb.close()
    
    return output_path